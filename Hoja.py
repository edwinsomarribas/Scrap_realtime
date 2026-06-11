import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime

# =========================
# Configuración del Excel
# =========================
NOMBRE_EXCEL = "datos_productividad.xlsx"
NOMBRE_HOJA = "Scrap"


# =========================
# FUNCIÓN: Guardar en Excel
# =========================
def guardar_en_excel(datos):

    ruta_excel = Path(NOMBRE_EXCEL)

    # Si el archivo existe, se abre. Si no, se crea
    if ruta_excel.exists():
        libro = load_workbook(NOMBRE_EXCEL)
    else:
        libro = Workbook()

    # Seleccionar o crear hoja
    if NOMBRE_HOJA in libro.sheetnames:
        hoja = libro[NOMBRE_HOJA]
    else:
        hoja = libro.create_sheet(NOMBRE_HOJA)

    # Eliminar hoja default si existe
    if "Sheet" in libro.sheetnames and len(libro.sheetnames) > 1:
        libro.remove(libro["Sheet"])

    # Crear encabezados si está vacío
    if hoja.max_row == 1 and hoja["A1"].value is None:
        hoja["A1"] = "Fecha"
        hoja["B1"] = "Producto"
        hoja["C1"] = "Job"
        hoja["D1"] = "Máquina"
        hoja["E1"] = "Scrap"
        hoja["F1"] = "Cantidad Scrap"
        hoja["G1"] = "Comentarios"

    # Siguiente fila disponible
    siguiente_fila = hoja.max_row + 1

    # Guardar datos
    hoja[f"A{siguiente_fila}"] = datos[0]  # Fecha
    hoja[f"B{siguiente_fila}"] = datos[1]  # Producto
    hoja[f"C{siguiente_fila}"] = datos[2]  # Job
    hoja[f"D{siguiente_fila}"] = datos[3]  # Máquina
    hoja[f"E{siguiente_fila}"] = datos[4]  # Scrap
    hoja[f"F{siguiente_fila}"] = datos[5]  # Cantidad Scrap
    hoja[f"G{siguiente_fila}"] = datos[6]  # Comentarios

    # Guardar archivo
    libro.save(NOMBRE_EXCEL)

    return siguiente_fila


# =========================
# FUNCIÓN: Actualizar máquinas según producto
# =========================
def actualizar_maquinas(event=None):

    producto = entrada_producto.get()

    if producto == "Varsity":
        opciones_maquina = ["261056"]

    elif producto == "Body":
        opciones_maquina = ["CR-261056", "CR-261057", "261056"]

    elif producto == "RO":
        opciones_maquina = ["Celda1", "Celda2", "CW1", "CW2", "CW3", "CW4"]

    elif producto == "Stryker":
        opciones_maquina = ["CW1", "CW2", "CW3", "CW4"]

    else:
        opciones_maquina = []

    entrada_maquina["values"] = opciones_maquina
    entrada_maquina.set("Seleccione una máquina")


# =========================
# FUNCIÓN: Limpiar campos después de guardar
# =========================
def limpiar_campos_despues_guardar():

    entrada_producto.set("Seleccione un producto")

    # IMPORTANTE:
    # No se borra el campo Job después de guardar
    # entrada_job.delete(0, tk.END)

    entrada_maquina["values"] = []
    entrada_maquina.set("Seleccione una máquina")

    entrada_scrap.set("Seleccione scrap")
    entrada_cantidad_scrap.delete(0, tk.END)
    entrada_comentarios.delete(0, tk.END)

    entrada_producto.focus()


# =========================
# FUNCIÓN: Limpiar todos los campos con botón Limpiar
# =========================
def limpiar_todos_los_campos():

    entrada_producto.set("Seleccione un producto")
    entrada_job.delete(0, tk.END)

    entrada_maquina["values"] = []
    entrada_maquina.set("Seleccione una máquina")

    entrada_scrap.set("Seleccione scrap")
    entrada_cantidad_scrap.delete(0, tk.END)
    entrada_comentarios.delete(0, tk.END)

    entrada_producto.focus()


# =========================
# FUNCIÓN: Botón Guardar
# =========================
def boton_guardar():

    producto = entrada_producto.get()
    job = entrada_job.get().strip()
    maquina = entrada_maquina.get()
    scrap = entrada_scrap.get()
    cantidad_scrap = entrada_cantidad_scrap.get().strip()
    comentarios = entrada_comentarios.get().strip()

    # Evitar guardar los textos iniciales de los menús
    if producto == "Seleccione un producto":
        producto = ""

    if maquina == "Seleccione una máquina":
        maquina = ""

    if scrap == "Seleccione scrap":
        scrap = ""

    # =========================
    # CHECKPOINT: Validar que todos los campos estén llenos
    # =========================
    campos_faltantes = []

    if producto == "":
        campos_faltantes.append("Producto")

    if job == "":
        campos_faltantes.append("Job")

    if maquina == "":
        campos_faltantes.append("Máquina")

    if scrap == "":
        campos_faltantes.append("Scrap")

    if cantidad_scrap == "":
        campos_faltantes.append("Cantidad Scrap")

    if comentarios == "":
        campos_faltantes.append("Comentarios")

    # Si hay campos vacíos, detener el guardado
    if campos_faltantes:
        mensaje = "Debe completar todos los campos antes de guardar.\n\n"
        mensaje += "Campos faltantes:\n"
        mensaje += "\n".join(f"- {campo}" for campo in campos_faltantes)

        messagebox.showwarning("Campos incompletos", mensaje)
        return

    # =========================
    # CHECKPOINT: Validar que Cantidad Scrap sea número
    # =========================
    try:
        cantidad_scrap_numero = int(cantidad_scrap)

        if cantidad_scrap_numero <= 0:
            messagebox.showwarning(
                "Cantidad inválida",
                "La cantidad de scrap debe ser mayor que 0."
            )
            return

    except ValueError:
        messagebox.showwarning(
            "Cantidad inválida",
            "La cantidad de scrap debe ser un número entero."
        )
        return

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")

    datos = [
        fecha,
        producto,
        job,
        maquina,
        scrap,
        cantidad_scrap_numero,
        comentarios
    ]

    try:
        fila_guardada = guardar_en_excel(datos)

        messagebox.showinfo(
            "Guardado exitoso",
            f"Datos guardados en la fila {fila_guardada}."
        )

        # Limpia campos después de guardar, pero mantiene Job
        limpiar_campos_despues_guardar()

    except PermissionError:
        messagebox.showerror(
            "Archivo abierto",
            "No se pudo guardar porque el archivo Excel está abierto.\n\n"
            "Cierre el archivo datos_productividad.xlsx e intente de nuevo."
        )

    except Exception as error:
        messagebox.showerror("Error", f"Ocurrió un error:\n{error}")


# =========================
# VENTANA PRINCIPAL
# =========================
ventana = tk.Tk()
ventana.title("Registro de Productividad Coils")
ventana.geometry("460x430")
ventana.resizable(True, True)


# =========================
# TÍTULO
# =========================
titulo = tk.Label(
    ventana,
    text="Ingreso de Scrap",
    font=("Arial", 16, "bold")
)
titulo.pack(pady=10)


# =========================
# FRAME FORMULARIO
# =========================
frame = tk.Frame(ventana)
frame.pack(pady=10)


# =========================
# Campo Producto - Menú desplegable
# =========================
tk.Label(frame, text="Producto:").grid(row=0, column=0, padx=10, pady=5, sticky="e")

entrada_producto = ttk.Combobox(
    frame,
    values=["Varsity", "Stryker", "Body", "RO"],
    width=27,
    state="readonly"
)
entrada_producto.grid(row=0, column=1)
entrada_producto.set("Seleccione un producto")

# Evento reactivo: cuando cambia Producto, cambia Máquina
entrada_producto.bind("<<ComboboxSelected>>", actualizar_maquinas)


# =========================
# Campo Job
# =========================
tk.Label(frame, text="Job:").grid(row=1, column=0, padx=10, pady=5, sticky="e")

entrada_job = tk.Entry(frame, width=30)
entrada_job.grid(row=1, column=1)


# =========================
# Campo Máquina - Menú desplegable reactivo
# =========================
tk.Label(frame, text="Máquina:").grid(row=2, column=0, padx=10, pady=5, sticky="e")

entrada_maquina = ttk.Combobox(
    frame,
    values=[],
    width=27,
    state="readonly"
)
entrada_maquina.grid(row=2, column=1)
entrada_maquina.set("Seleccione una máquina")


# =========================
# Campo Scrap - Menú desplegable
# =========================
tk.Label(frame, text="Scrap:").grid(row=3, column=0, padx=10, pady=5, sticky="e")

entrada_scrap = ttk.Combobox(
    frame,
    values=["C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8", "C9"],
    width=27,
    state="readonly"
)
entrada_scrap.grid(row=3, column=1)
entrada_scrap.set("Seleccione scrap")


# =========================
# Campo Cantidad Scrap
# =========================
tk.Label(frame, text="Cantidad Scrap:").grid(row=4, column=0, padx=10, pady=5, sticky="e")

entrada_cantidad_scrap = tk.Entry(frame, width=30)
entrada_cantidad_scrap.grid(row=4, column=1)


# =========================
# Campo Comentarios
# =========================
tk.Label(frame, text="Comentarios:").grid(row=5, column=0, padx=10, pady=5, sticky="e")

entrada_comentarios = tk.Entry(frame, width=30)
entrada_comentarios.grid(row=5, column=1)


# =========================
# FRAME BOTONES
# =========================
frame_botones = tk.Frame(ventana)
frame_botones.pack(pady=15)


# Botón guardar
btn_guardar = tk.Button(
    frame_botones,
    text="Guardar",
    command=boton_guardar,
    bg="#4CAF50",
    fg="white",
    width=12
)
btn_guardar.grid(row=0, column=0, padx=5)


# Botón limpiar
btn_limpiar = tk.Button(
    frame_botones,
    text="Limpiar",
    command=limpiar_todos_los_campos,
    bg="orange",
    width=12
)
btn_limpiar.grid(row=0, column=1, padx=5)


# Botón salir
btn_salir = tk.Button(
    frame_botones,
    text="Salir",
    command=ventana.destroy,
    bg="red",
    fg="white",
    width=12
)
btn_salir.grid(row=0, column=2, padx=5)


# =========================
# TEXTO INFERIOR
# =========================
nota = tk.Label(
    ventana,
    text=f"Archivo: {NOMBRE_EXCEL}",
    font=("Arial", 9)
)
nota.pack(pady=5)


# Cursor inicial
entrada_producto.focus()

# Ejecutar app
ventana.mainloop()