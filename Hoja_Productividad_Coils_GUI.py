import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime

# =========================
# Configuración del Excel
# =========================

NOMBRE_EXCEL = "datos_productividad.xlsx"

# Para Android (Pydroid) puedes usar:
# NOMBRE_EXCEL = "/storage/emulated/0/Download/datos_productividad.xlsx"

NOMBRE_HOJA = "Scrap"


# =========================
# FUNCIÓN: Guardar en Excel
# =========================

def guardar_en_excel(datos):

    ruta_excel = Path(NOMBRE_EXCEL)

    if ruta_excel.exists():
        libro = load_workbook(NOMBRE_EXCEL)
    else:
        libro = Workbook()

    if NOMBRE_HOJA in libro.sheetnames:
        hoja = libro[NOMBRE_HOJA]
    else:
        hoja = libro.create_sheet(NOMBRE_HOJA)

    if "Sheet" in libro.sheetnames and len(libro.sheetnames) > 1:
        libro.remove(libro["Sheet"])

    if hoja.max_row == 1 and hoja["A1"].value is None:
        hoja.append([
            "Fecha",
            "Producto",
            "Job",
            "Máquina",
            "Scrap",
            "Cantidad Scrap",
            "Comentarios",
            "Operario"
        ])

    hoja.append(datos)

    fila_guardada = hoja.max_row

    libro.save(NOMBRE_EXCEL)

    return fila_guardada


# =========================
# FUNCIÓN: Actualizar máquinas
# =========================

def actualizar_maquinas(event=None):

    producto = entrada_producto.get()

    if producto == "Varsity":
        opciones_maquina = ["261056"]

    elif producto == "Body":
        opciones_maquina = [
            "CR-261056",
            "CR-261057",
            "261056"
        ]

    elif producto == "RO":
        opciones_maquina = [
            "Celda1",
            "Celda2",
            "CW1",
            "CW2",
            "CW3",
            "CW4"
        ]

    elif producto == "Stryker":
        opciones_maquina = [
            "CW1",
            "CW2",
            "CW3",
            "CW4"
        ]

    else:
        opciones_maquina = []

    entrada_maquina["values"] = opciones_maquina
    entrada_maquina.set("Seleccione una máquina")


# =========================
# FUNCIÓN: Limpiar todo
# =========================

def limpiar_campos():

    entrada_producto.set("Seleccione un producto")
    entrada_job.delete(0, tk.END)

    entrada_maquina["values"] = []
    entrada_maquina.set("Seleccione una máquina")

    entrada_scrap.set("Seleccione scrap")

    entrada_cantidad_scrap.delete(0, tk.END)
    entrada_comentarios.delete(0, tk.END)
    entrada_operarios.delete(0, tk.END)

    entrada_producto.focus()


# =========================
# FUNCIÓN: Limpiar después guardar
# =========================

def limpiar_despues_de_guardar():

    entrada_scrap.set("Seleccione scrap")
    entrada_cantidad_scrap.delete(0, tk.END)
    entrada_comentarios.delete(0, tk.END)
    entrada_operarios.delete(0, tk.END)

    entrada_scrap.focus()


# =========================
# FUNCIÓN: Guardar
# =========================

def boton_guardar():

    producto = entrada_producto.get()
    job = entrada_job.get().strip()
    maquina = entrada_maquina.get()
    scrap = entrada_scrap.get()
    cantidad_scrap = entrada_cantidad_scrap.get().strip()
    comentarios = entrada_comentarios.get().strip()
    operarios = entrada_operarios.get().strip()

    if producto == "Seleccione un producto":
        producto = ""

    if maquina == "Seleccione una máquina":
        maquina = ""

    if scrap == "Seleccione scrap":
        scrap = ""

    campos_faltantes = []

    if not producto:
        campos_faltantes.append("Producto")

    if not job:
        campos_faltantes.append("Job")

    if not maquina:
        campos_faltantes.append("Máquina")

    if not scrap:
        campos_faltantes.append("Scrap")

    if not cantidad_scrap:
        campos_faltantes.append("Cantidad Scrap")

    if not comentarios:
        campos_faltantes.append("Comentarios")

    if not operarios:
        campos_faltantes.append("Operarios")

    if campos_faltantes:

        mensaje = "Debe completar todos los campos antes de guardar.\n\n"
        mensaje += "Campos faltantes:\n"
        mensaje += "\n".join(f"• {campo}" for campo in campos_faltantes)

        messagebox.showwarning(
            "Campos incompletos",
            mensaje
        )

        return

    try:

        cantidad_scrap_numero = int(cantidad_scrap)

        if cantidad_scrap_numero <= 0:

            messagebox.showwarning(
                "Cantidad inválida",
                "La cantidad de scrap debe ser mayor que cero."
            )

            return

    except ValueError:

        messagebox.showwarning(
            "Cantidad inválida",
            "La cantidad de scrap debe ser un número entero."
        )

        return

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")

    datos = [
        fecha,
        producto,
        job,
        maquina,
        scrap,
        cantidad_scrap_numero,
        comentarios,
        operarios
    ]

    try:

        fila_guardada = guardar_en_excel(datos)

        messagebox.showinfo(
            "Guardado exitoso",
            f"Datos guardados en la fila {fila_guardada}."
        )

        limpiar_despues_de_guardar()

    except PermissionError:

        messagebox.showerror(
            "Archivo abierto",
            "No se pudo guardar porque el archivo Excel está abierto.\n\n"
            "Cierre el archivo y vuelva a intentarlo."
        )

    except Exception as error:

        messagebox.showerror(
            "Error",
            f"Ocurrió un error:\n{error}"
        )


# =========================
# VENTANA PRINCIPAL
# =========================

ventana = tk.Tk()

ventana.title("Registro de Productividad Coils")
ventana.geometry("500x520")
ventana.resizable(True, True)


# =========================
# TÍTULO
# =========================

titulo = tk.Label(
    ventana,
    text="Ingreso de Scrap",
    font=("Arial", 16, "bold")
)

titulo.pack(pady=10)


# =========================
# FRAME FORMULARIO
# =========================

frame = tk.Frame(ventana)
frame.pack(pady=10)


# =========================
# PRODUCTO
# =========================

tk.Label(
    frame,
    text="Producto:"
).grid(
    row=0,
    column=0,
    padx=10,
    pady=5,
    sticky="e"
)

entrada_producto = ttk.Combobox(
    frame,
    values=["Varsity", "Stryker", "Body", "RO"],
    width=27,
    state="readonly"
)

entrada_producto.grid(row=0, column=1)
entrada_producto.set("Seleccione un producto")

entrada_producto.bind(
    "<<ComboboxSelected>>",
    actualizar_maquinas
)


# =========================
# JOB
# =========================

tk.Label(
    frame,
    text="Job:"
).grid(
    row=1,
    column=0,
    padx=10,
    pady=5,
    sticky="e"
)

entrada_job = tk.Entry(frame, width=30)
entrada_job.grid(row=1, column=1)


# =========================
# MÁQUINA
# =========================

tk.Label(
    frame,
    text="Máquina:"
).grid(
    row=2,
    column=0,
    padx=10,
    pady=5,
    sticky="e"
)

entrada_maquina = ttk.Combobox(
    frame,
    values=[],
    width=27,
    state="readonly"
)

entrada_maquina.grid(row=2, column=1)
entrada_maquina.set("Seleccione una máquina")


# =========================
# SCRAP
# =========================

tk.Label(
    frame,
    text="Scrap:"
).grid(
    row=3,
    column=0,
    padx=10,
    pady=5,
    sticky="e"
)

entrada_scrap = ttk.Combobox(
    frame,
    values=["C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8", "C9"],
    width=27,
    state="readonly"
)

entrada_scrap.grid(row=3, column=1)
entrada_scrap.set("Seleccione scrap")


# =========================
# CANTIDAD SCRAP
# =========================

tk.Label(
    frame,
    text="Cantidad Scrap:"
).grid(
    row=4,
    column=0,
    padx=10,
    pady=5,
    sticky="e"
)

entrada_cantidad_scrap = tk.Entry(frame, width=30)
entrada_cantidad_scrap.grid(row=4, column=1)


# =========================
# COMENTARIOS
# =========================

tk.Label(
    frame,
    text="Comentarios:"
).grid(
    row=5,
    column=0,
    padx=10,
    pady=5,
    sticky="e"
)

entrada_comentarios = tk.Entry(frame, width=30)
entrada_comentarios.grid(row=5, column=1)


# =========================
# OPERARIOS
# =========================

tk.Label(
    frame,
    text="Operarios:"
).grid(
    row=6,
    column=0,
    padx=10,
    pady=5,
    sticky="e"
)

entrada_operarios = tk.Entry(frame, width=30)
entrada_operarios.grid(row=6, column=1)


# =========================
# BOTONES
# =========================

frame_botones = tk.Frame(ventana)
frame_botones.pack(pady=15)

btn_guardar = tk.Button(
    frame_botones,
    text="Guardar",
    command=boton_guardar,
    bg="#4CAF50",
    fg="white",
    width=12
)

btn_guardar.grid(row=0, column=0, padx=5)

btn_limpiar = tk.Button(
    frame_botones,
    text="Limpiar",
    command=limpiar_campos,
    bg="orange",
    width=12
)

btn_limpiar.grid(row=0, column=1, padx=5)

btn_salir = tk.Button(
    frame_botones,
    text="Salir",
    command=ventana.destroy,
    bg="red",
    fg="white",
    width=12
)

btn_salir.grid(row=0, column=2, padx=5)


# =========================
# NOTA INFERIOR
# =========================

nota = tk.Label(
    ventana,
    text=f"Archivo: {NOMBRE_EXCEL}",
    font=("Arial", 9)
)

nota.pack(pady=5)


# =========================
# INICIO
# =========================

entrada_producto.focus()

ventana.mainloop()