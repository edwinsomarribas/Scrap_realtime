from openpyxl import Workbook, load_workbook
from pathlib import Path


# Nombre del archivo Excel
nombre_excel = "datos_productividad.xlsx"

# Nombre de la hoja
nombre_hoja = "Scrap"

# Crear ruta del archivo
ruta_excel = Path(nombre_excel)


# Verificar si el archivo Excel ya existe
if ruta_excel.exists():
    libro = load_workbook(nombre_excel)
else:
    libro = Workbook()


# Verificar si la hoja existe
if nombre_hoja in libro.sheetnames:
    hoja = libro[nombre_hoja]
else:
    hoja = libro.create_sheet(nombre_hoja)


# Si el archivo es nuevo y tiene la hoja por defecto "Sheet", opcionalmente la eliminamos
if "Sheet" in libro.sheetnames and "Datos" in libro.sheetnames and len(libro.sheetnames) > 1:
    hoja_default = libro["Sheet"]
    libro.remove(hoja_default)


# Crear encabezados si la hoja está vacía
if hoja.max_row == 1 and hoja["A1"].value is None:
    hoja["A1"] = "Fecha"
    hoja["B1"] = "Producto"
    hoja["C1"] = "Job"
    hoja["D1"] = "Máquina"
    hoja["E1"] = "Scrap"
    hoja["F1"] = "Comentarios   "


# Buscar la siguiente fila vacía
siguiente_fila = hoja.max_row + 1


# Solicitar datos al usuario
dato_a = input("Ingrese el dato para la columna A: ")
dato_b = input("Ingrese el dato para la columna B: ")
dato_c = input("Ingrese el dato para la columna C: ")
dato_d = input("Ingrese el dato para la columna D: ")
dato_e = input("Ingrese el dato para la columna E: ")
dato_f   = input("Ingrese el dato para la columna F: ")


# Guardar los datos en la siguiente fila disponible
hoja[f"A{siguiente_fila}"] = dato_a
hoja[f"B{siguiente_fila}"] = dato_b
hoja[f"C{siguiente_fila}"] = dato_c
hoja[f"D{siguiente_fila}"] = dato_d
hoja[f"E{siguiente_fila}"] = dato_e


# Guardar el archivo Excel
libro.save(nombre_excel)

print("Datos guardados correctamente.")
print(f"Archivo actualizado: {nombre_excel}")
print(f"Datos guardados en la fila: {siguiente_fila}")