from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.popup import Popup


NOMBRE_HOJA = "Scrap"


class ScrapApp(App):

    def build(self):
        self.title = "Registro de Scrap"

        # Guardar el Excel en la carpeta interna de la app.
        # Esto evita problemas de permisos en Android.
        self.nombre_excel = str(Path(self.user_data_dir) / "datos_productividad.xlsx")

        layout_principal = BoxLayout(
            orientation="vertical",
            padding=20,
            spacing=15
        )

        titulo = Label(
            text="Ingreso de Scrap",
            font_size="24sp",
            size_hint=(1, 0.15)
        )
        layout_principal.add_widget(titulo)

        formulario = GridLayout(
            cols=2,
            spacing=10,
            size_hint=(1, 0.65)
        )

        # =========================
        # Producto
        # =========================
        formulario.add_widget(Label(text="Producto:", font_size="18sp"))

        self.producto = Spinner(
            text="Seleccione un producto",
            values=["Varsity", "Stryker", "Body", "RO"],
            font_size="18sp"
        )
        self.producto.bind(text=self.actualizar_maquinas)
        formulario.add_widget(self.producto)

        # =========================
        # Job
        # =========================
        formulario.add_widget(Label(text="Job:", font_size="18sp"))

        self.job = TextInput(
            multiline=False,
            font_size="18sp"
        )
        formulario.add_widget(self.job)

        # =========================
        # Máquina
        # =========================
        formulario.add_widget(Label(text="Máquina:", font_size="18sp"))

        self.maquina = Spinner(
            text="Seleccione una máquina",
            values=[],
            font_size="18sp"
        )
        formulario.add_widget(self.maquina)

        # =========================
        # Scrap
        # =========================
        formulario.add_widget(Label(text="Scrap:", font_size="18sp"))

        self.scrap = Spinner(
            text="Seleccione scrap",
            values=["C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8", "C9"],
            font_size="18sp"
        )
        formulario.add_widget(self.scrap)

        # =========================
        # Cantidad Scrap
        # =========================
        formulario.add_widget(Label(text="Cantidad Scrap:", font_size="18sp"))

        self.cantidad_scrap = TextInput(
            multiline=False,
            input_filter="int",
            font_size="18sp"
        )
        formulario.add_widget(self.cantidad_scrap)

        # =========================
        # Comentarios
        # =========================
        formulario.add_widget(Label(text="Comentarios:", font_size="18sp"))

        self.comentarios = TextInput(
            multiline=False,
            font_size="18sp"
        )
        formulario.add_widget(self.comentarios)

        layout_principal.add_widget(formulario)

        # =========================
        # Botones
        # =========================
        botones = BoxLayout(
            orientation="horizontal",
            spacing=10,
            size_hint=(1, 0.15)
        )

        btn_guardar = Button(
            text="Guardar",
            font_size="18sp",
            background_color=(0.2, 0.7, 0.3, 1)
        )
        btn_guardar.bind(on_press=self.boton_guardar)
        botones.add_widget(btn_guardar)

        btn_limpiar = Button(
            text="Limpiar",
            font_size="18sp",
            background_color=(1, 0.6, 0.1, 1)
        )
        btn_limpiar.bind(on_press=self.limpiar_todo)
        botones.add_widget(btn_limpiar)

        btn_salir = Button(
            text="Salir",
            font_size="18sp",
            background_color=(0.8, 0.1, 0.1, 1)
        )
        btn_salir.bind(on_press=self.salir_app)
        botones.add_widget(btn_salir)

        layout_principal.add_widget(botones)

        ruta_label = Label(
            text=f"Archivo: {self.nombre_excel}",
            font_size="12sp",
            size_hint=(1, 0.08)
        )
        layout_principal.add_widget(ruta_label)

        return layout_principal

    # =========================
    # Actualizar máquinas según producto
    # =========================
    def actualizar_maquinas(self, spinner, producto):

        if producto == "Varsity":
            opciones_maquina = ["261056"]

        elif producto == "Body":
            opciones_maquina = ["CR-261056", "CR-261057", "261056"]

        elif producto == "RO":
            opciones_maquina = ["Celda1", "Celda2", "CW1", "CW2", "CW3", "CW4"]

        elif producto == "Stryker":
            opciones_maquina = ["CW1", "CW2", "CW3", "CW4"]

        else:
            opciones_maquina = []

        self.maquina.values = opciones_maquina
        self.maquina.text = "Seleccione una máquina"

    # =========================
    # Guardar en Excel
    # =========================
    def guardar_en_excel(self, datos):

        ruta_excel = Path(self.nombre_excel)

        if ruta_excel.exists():
            libro = load_workbook(self.nombre_excel)
        else:
            libro = Workbook()

        if NOMBRE_HOJA in libro.sheetnames:
            hoja = libro[NOMBRE_HOJA]
        else:
            hoja = libro.create_sheet(NOMBRE_HOJA)

        if "Sheet" in libro.sheetnames and len(libro.sheetnames) > 1:
            libro.remove(libro["Sheet"])

        if hoja.max_row == 1 and hoja["A1"].value is None:
            hoja["A1"] = "Fecha"
            hoja["B1"] = "Producto"
            hoja["C1"] = "Job"
            hoja["D1"] = "Máquina"
            hoja["E1"] = "Scrap"
            hoja["F1"] = "Cantidad Scrap"
            hoja["G1"] = "Comentarios"

        siguiente_fila = hoja.max_row + 1

        hoja[f"A{siguiente_fila}"] = datos[0]
        hoja[f"B{siguiente_fila}"] = datos[1]
        hoja[f"C{siguiente_fila}"] = datos[2]
        hoja[f"D{siguiente_fila}"] = datos[3]
        hoja[f"E{siguiente_fila}"] = datos[4]
        hoja[f"F{siguiente_fila}"] = datos[5]
        hoja[f"G{siguiente_fila}"] = datos[6]

        libro.save(self.nombre_excel)

        return siguiente_fila

    # =========================
    # Botón Guardar
    # =========================
    def boton_guardar(self, instance):

        producto = self.producto.text
        job = self.job.text.strip()
        maquina = self.maquina.text
        scrap = self.scrap.text
        cantidad_scrap = self.cantidad_scrap.text.strip()
        comentarios = self.comentarios.text.strip()

        if producto == "Seleccione un producto":
            producto = ""

        if maquina == "Seleccione una máquina":
            maquina = ""

        if scrap == "Seleccione scrap":
            scrap = ""

        campos_faltantes = []

        if producto == "":
            campos_faltantes.append("Producto")

        if job == "":
            campos_faltantes.append("Job")

        if maquina == "":
            campos_faltantes.append("Máquina")

        if scrap == "":
            campos_faltantes.append("Scrap")

        if cantidad_scrap == "":
            campos_faltantes.append("Cantidad Scrap")

        if comentarios == "":
            campos_faltantes.append("Comentarios")

        if campos_faltantes:
            mensaje = "Debe completar todos los campos antes de guardar.\n\n"
            mensaje += "Campos faltantes:\n"
            mensaje += "\n".join(f"- {campo}" for campo in campos_faltantes)
            self.mostrar_popup("Campos incompletos", mensaje)
            return

        try:
            cantidad_scrap_numero = int(cantidad_scrap)

            if cantidad_scrap_numero <= 0:
                self.mostrar_popup(
                    "Cantidad inválida",
                    "La cantidad de scrap debe ser mayor que 0."
                )
                return

        except ValueError:
            self.mostrar_popup(
                "Cantidad inválida",
                "La cantidad de scrap debe ser un número entero."
            )
            return

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M")

        datos = [
            fecha,
            producto,
            job,
            maquina,
            scrap,
            cantidad_scrap_numero,
            comentarios
        ]

        try:
            fila_guardada = self.guardar_en_excel(datos)

            self.mostrar_popup(
                "Guardado exitoso",
                f"Datos guardados en la fila {fila_guardada}.\n\nArchivo:\n{self.nombre_excel}"
            )

            # No borrar Producto, Job ni Máquina.
            self.limpiar_despues_de_guardar()

        except Exception as error:
            self.mostrar_popup(
                "Error",
                f"Ocurrió un error al guardar:\n{error}"
            )

    # =========================
    # Limpiar después de guardar
    # =========================
    def limpiar_despues_de_guardar(self):

        self.scrap.text = "Seleccione scrap"
        self.cantidad_scrap.text = ""
        self.comentarios.text = ""

    # =========================
    # Limpiar todo manualmente
    # =========================
    def limpiar_todo(self, instance):

        self.producto.text = "Seleccione un producto"
        self.job.text = ""

        self.maquina.values = []
        self.maquina.text = "Seleccione una máquina"

        self.scrap.text = "Seleccione scrap"
        self.cantidad_scrap.text = ""
        self.comentarios.text = ""

    # =========================
    # Popup
    # =========================
    def mostrar_popup(self, titulo, mensaje):

        contenido = BoxLayout(
            orientation="vertical",
            padding=15,
            spacing=15
        )

        contenido.add_widget(Label(text=mensaje))

        boton_cerrar = Button(
            text="OK",
            size_hint=(1, 0.3)
        )

        popup = Popup(
            title=titulo,
            content=contenido,
            size_hint=(0.85, 0.55)
        )

        boton_cerrar.bind(on_press=popup.dismiss)
        contenido.add_widget(boton_cerrar)

        popup.open()

    # =========================
    # Salir
    # =========================
    def salir_app(self, instance):
        App.get_running_app().stop()


if __name__ == "__main__":
    ScrapApp().run()